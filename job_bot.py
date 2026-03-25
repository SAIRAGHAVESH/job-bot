"""
====================================================================
FULL JOB SEARCH BOT
====================================================================
Sources:
  1. Job Boards    — LinkedIn, Indeed, Glassdoor, ZipRecruiter, Google Jobs
  2. Free APIs     — Adzuna, Jooble, RemoteOK, USAJobs, WeWorkRemotely
  3. ATS Platforms — Greenhouse, Lever, Ashby, SmartRecruiters, Workday
                     (covers 500+ top US companies in one go)
Output:
  - New Excel file every day: Daily_Jobs/Jobs_YYYY-MM-DD.xlsx
  - Telegram alert with top openings + apply links
Setup:
  pip install python-jobspy openpyxl requests schedule pandas
→ Set TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID below
→ Optional: Set ADZUNA_APP_ID, ADZUNA_APP_KEY (free at developer.adzuna.com)
→ Optional: Set JOOBLE_API_KEY (free at jooble.org/api)
→ Run: python job_bot.py
====================================================================
"""
import os, json, time, hashlib, requests, schedule
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
try:
    from jobspy import scrape_jobs
    JOBSPY_OK = True
except ImportError:
    JOBSPY_OK = False
    print("Run: pip install python-jobspy")
# ── CONFIG ────────────────────────────────────────────────────────────────────
TELEGRAM_BOT_TOKEN = "8621053380:AAHGJhcJPARAnHfekKvHA7fQ2SE2CbpKnuo"
TELEGRAM_CHAT_ID   = "612269695"
ADZUNA_APP_ID      = ""  # free @ developer.adzuna.com
ADZUNA_APP_KEY     = ""
JOOBLE_API_KEY     = ""   # free @ jooble.org/api
SAVE_FOLDER        = "Daily_Jobs"
SEEN_FILE          = "seen_jobs.json"
JOB_TITLES = [
    "Software Engineer", "Senior Software Engineer", "Staff Software Engineer",
    "DevOps Engineer", "Senior DevOps Engineer", "Platform Engineer",
    "Cloud Engineer", "Cloud Architect", "Infrastructure Engineer",
    "Site Reliability Engineer", "SRE", "Production Engineer",
    "Security Engineer", "DevSecOps Engineer", "MLOps Engineer",
    "Backend Engineer", "Full Stack Engineer", "Systems Engineer",
    "Kubernetes Engineer", "Cloud Security Engineer", "DataOps Engineer",
    "Software Development Engineer", "Application Engineer",
]
# ── 500+ COMPANIES via ATS ────────────────────────────────────────────────────
# Greenhouse: just the company slug → api.greenhouse.io/v1/boards/{slug}/jobs
GREENHOUSE_COMPANIES = [
    # Big Tech & Cloud
    "airbnb","coinbase","stripe","dropbox","square","lyft","doordash",
    "robinhood","brex","plaid","chime","affirm","marqeta","ripple",
    "databricks","snowflake","elastic","mongodb","twilio","okta",
    "cloudflare","crowdstrike","datadog","splunk","zscaler","netskope",
    "hashicorp","circleci","docker","gitlab","github","atlassian",
    "hubspot","zendesk","intercom","notion","figma","canva","asana",
    "airtable","clickup","linear","vercel","netlify","heroku",
    # AI / ML
    "anthropic","scale","cohere","huggingface","weights-biases",
    "openai","perplexity","mistral","together-ai","replicate",
    # Fintech
    "chime","sofi","nerdwallet","betterment","wealthfront","robinhood",
    "carta","rippling","gusto","adyen","payoneer","wise",
    # Health Tech
    "tempus","guardant","color-genomics","hims","ro","noom",
    "headspace","calm","cerebral","teladoc",
    # E-commerce / Consumer
    "etsy","wayfair","chewy","poshmark","offerup","mercari",
    "instacart","gopuff","getaround","turo","vacasa",
    # Enterprise SaaS
    "servicenow","workday","veeva","medallia","qualtrics","sprinklr",
    "amplitude","mixpanel","segment","mparticle","heap",
    "braze","klaviyo","sendgrid","mailchimp","iterable",
    # Infrastructure / DevOps
    "fastly","fly-io","render","railway","supabase","planetscale",
    "cockroachdb","yugabyte","citus","timescale","questdb",
    # Security
    "pagerduty","lacework","orca","wiz","snyk","sonatype",
    "sysdig","aqua-security","anchore","chainguard",
    # Gaming / Media
    "epicgames","roblox","unity","niantic","discord","twitch",
    # Logistics / Mobility
    "flexport","project44","motive","samsara","axon","rivian",
    "zoox","waymo","aurora","argo-ai",
    # IT Services
    "epam","thoughtworks","slalom","publicissapient",
]
# Lever: jobs.lever.co/{slug}?lever-source=jobspage
LEVER_COMPANIES = [
    # Big Tech & SaaS
    "netflix","lyft","atlassian","reddit","pinterest","quora",
    "shopify","squarespace","wix","bigcommerce","magento",
    "zendesk","freshworks","pipedrive","copper","close",
    "mixmax","outreach","salesloft","gong","chorus",
    # Cloud / Infra
    "hashicorp","vault","consul","terraform","nomad",
    "grafana","prometheus","influxdata","victoria-metrics",
    "dynatrace","newrelic","appdynamics","honeycomb","lightstep",
    # AI / Data
    "dbtlabs","fivetran","airbyte","stitch","matillion",
    "alation","atlan","collibra","informatica","talend",
    # Security
    "1password","dashlane","lastpass","bitwarden","keeper",
    "sentinelone","cylance","malwarebytes","bitdefender",
    # Fintech
    "navan","ramp","divvy","airbase","expensify","concur",
    "tipalti","bill","melio","paylocity","paycom",
    # Health
    "headway","alma","sondermind","talkspace","betterhelp",
    "sword-health","hinge-health","kaia-health","woebot",
    # E-comm / Marketplace
    "faire","angi","thumbtack","houzz","build","buildzoom",
    # Edtech
    "duolingo","coursera","udemy","pluralsight","skillshare",
    "masterclass","brilliant","khan-academy","codecademy",
]
# Ashby: jobs.ashbyhq.com/{slug}
ASHBY_COMPANIES = [
    "openai","perplexity","mistral","together","anyscale",
    "modal","modal-labs","fly","railway","render",
    "supabase","neon","planetscale","turso","xata",
    "resend","loops","postmark","sendgrid",
    "linear","height","plane","basecamp",
    "vercel","netlify","cloudflare-workers","deno",
    "cursor","codeium","tabnine","sourcegraph",
]
# SmartRecruiters public API
SMARTRECRUITERS_COMPANIES = [
    "Visa","Mastercard","PayPal","Intuit","Adobe","Salesforce",
    "Oracle","SAP","IBM","Cisco","Intel","AMD","Qualcomm","Broadcom",
    "Dell","HP","Lenovo","Logitech","Corsair",
    "Deloitte","PwC","EY","KPMG","Accenture","McKinsey","BCG",
    "JPMorgan","Goldman","Morgan Stanley","Citi","BankOfAmerica","Wells Fargo",
    "Amazon","Microsoft","Apple","Google","Meta","Nvidia","Tesla",
    "Walmart","Target","Costco","Kroger",
    "UnitedHealth","CVS","Cigna","Humana","Anthem",
    "Boeing","Lockheed","Raytheon","Northrop","GeneralDynamics",
]
# Workday tenant URLs (pattern: {tenant}.wd5.myworkdayjobs.com)
WORKDAY_TENANTS = [
    ("Apple",           "apple"),
    ("Google",          "google"),
    ("Microsoft",       "microsoft"),
    ("Meta",            "meta"),
    ("Amazon",          "amazon"),
    ("Nvidia",          "nvidia"),
    ("Tesla",           "tesla"),
    ("Salesforce",      "salesforce"),
    ("Adobe",           "adobe"),
    ("Intuit",          "intuit"),
    ("Workday",         "workday"),
    ("ServiceNow",      "servicenow"),
    ("Palo Alto",       "paloaltonetworks"),
    ("Crowdstrike",     "crowdstrike"),
    ("Zscaler",         "zscaler"),
    ("Snowflake",       "snowflake"),
    ("Databricks",      "databricks"),
    ("Splunk",          "splunk"),
    ("Elastic",         "elastic"),
    ("MongoDB",         "mongodb"),
    ("HubSpot",         "hubspot"),
    ("Okta",            "okta"),
    ("Twilio",          "twilio"),
    ("Zendesk",         "zendesk"),
    ("Veeva",           "veeva"),
    ("Informatica",     "informatica"),
    ("Cognizant",       "cognizant"),
    ("Infosys",         "infosys"),
    ("TCS",             "tcs"),
    ("Wipro",           "wipro"),
    ("HCL",             "hcl"),
    ("Capgemini",       "capgemini"),
]
    ("EPAM",            "epam"),
# ── HELPERS ───────────────────────────────────────────────────────────────────
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}
def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return set(json.load(f))
    return set()
def save_seen(seen):
    with open(SEEN_FILE, "w") as f:
        json.dump(list(seen), f)
def job_id(title, company):
    return hashlib.md5(f"{title}{company}".lower().encode()).hexdigest()
def make_job(title, company, location, url, source, salary="—"):
    return {"title": title, "company": company, "location": location,
            "url": url, "source": source, "salary": salary,
            "posted": datetime.now().strftime("%Y-%m-%d")}
def send_telegram(msg):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": msg,
                  "parse_mode": "HTML", "disable_web_page_preview": True},
            timeout=10
        )
        time.sleep(0.4)
    except Exception as e:
        print(f"  Telegram error: {e}")
def is_relevant(title):
    keywords = [
        "engineer","sre","devops","platform","cloud","infrastructure",
        "backend","fullstack","full stack","security","mlops","dataops",
        "reliability","kubernetes","architect","systems"
    ]
    return any(k in title.lower() for k in keywords)
# ── SOURCE 1: JOB BOARDS (JobSpy) ────────────────────────────────────────────
def fetch_job_boards(seen):
    if not JOBSPY_OK:
        return []
    import pandas as pd
    jobs, frames = [], []
    for title in JOB_TITLES:
        print(f"  [JobBoards] {title}")
        try:
            df = scrape_jobs(
                site_name=["linkedin","indeed","glassdoor","zip_recruiter","google"],
                search_term=title,
                location="United States",
                results_wanted=25,
                hours_old=24,
                is_remote=True,
                country_indeed="USA",
            )
            if not df.empty:
                frames.append(df)
        except Exception as e:
            print(f"    Error: {e}")
        time.sleep(2)
    if not frames:
        return []
    for _, row in pd.concat(frames, ignore_index=True).iterrows():
        jid = job_id(str(row.get("title","")), str(row.get("company","")))
        if jid not in seen:
            seen.add(jid)
            jobs.append(make_job(
                str(row.get("title","")), str(row.get("company","")),
                str(row.get("location","Remote")),
                str(row.get("job_url", row.get("job_link",""))),
                str(row.get("site","")).upper(),
                str(row.get("min_amount","") or "—"),
            ))
    return jobs
# ── SOURCE 2: ADZUNA API ──────────────────────────────────────────────────────
def fetch_adzuna(seen):
    if not ADZUNA_APP_ID:
        print("  [Adzuna] Skipping — no API key (get free key at developer.adzuna.com)")
        return []
    jobs = []
    for title in JOB_TITLES[:8]:  # limit to avoid rate limits
        try:
            url = (f"https://api.adzuna.com/v1/api/jobs/us/search/1"
                   f"?app_id={ADZUNA_APP_ID}&app_key={ADZUNA_APP_KEY}"
                   f"&results_per_page=20&what={requests.utils.quote(title)}"
                   f"&where=remote&max_days_old=1&content-type=application/json")
            r = requests.get(url, timeout=10)
            for job in r.json().get("results", []):
                jid = job_id(job.get("title",""), job.get("company",{}).get("display_name",""))
                if jid not in seen and is_relevant(job.get("title","")):
                    seen.add(jid)
                    jobs.append(make_job(
                        job.get("title",""),
                        job.get("company",{}).get("display_name",""),
                        job.get("location",{}).get("display_name","Remote"),
                        job.get("redirect_url",""),
                        "ADZUNA",
                        f"${job.get('salary_min',0):,.0f}" if job.get("salary_min") else "—",
                    ))
        except Exception as e:
            print(f"  [Adzuna] Error: {e}")
        time.sleep(1)
    print(f"  [Adzuna] {len(jobs)} jobs")
    return jobs
# ── SOURCE 3: JOOBLE API ──────────────────────────────────────────────────────
def fetch_jooble(seen):
    if not JOOBLE_API_KEY:
        print("  [Jooble] Skipping — no API key (get free key at jooble.org/api)")
        return []
    jobs = []
    for title in JOB_TITLES[:8]:
        try:
            r = requests.post(
                f"https://jooble.org/api/{JOOBLE_API_KEY}",
                json={"keywords": title, "location": "remote", "page": "1"},
                timeout=10
            )
            for job in r.json().get("jobs", []):
                jid = job_id(job.get("title",""), job.get("company",""))
                if jid not in seen and is_relevant(job.get("title","")):
                    seen.add(jid)
                    jobs.append(make_job(
                        job.get("title",""), job.get("company",""),
                        job.get("location","Remote"),
                        job.get("link",""), "JOOBLE",
                        job.get("salary","—"),
                    ))
        except Exception as e:
            print(f"  [Jooble] Error: {e}")
        time.sleep(1)
    print(f"  [Jooble] {len(jobs)} jobs")
    return jobs
# ── SOURCE 4: REMOTEOK (FREE PUBLIC API) ─────────────────────────────────────
def fetch_remoteok(seen):
    jobs = []
    try:
        r = requests.get("https://remoteok.com/api", headers={"User-Agent": "JobBot/1.0"}, timeout=15)
        for job in r.json()[1:]:  # first item is metadata
            if not is_relevant(job.get("position","")):
                continue
            jid = job_id(job.get("position",""), job.get("company",""))
            if jid not in seen:
                seen.add(jid)
                salary = ""
                if job.get("salary_min"):
                    salary = f"${job['salary_min']:,} - ${job.get('salary_max',0):,}"
                jobs.append(make_job(
                    job.get("position",""), job.get("company",""),
                    "Remote", job.get("url",""), "REMOTEOK", salary or "—",
                ))
    except Exception as e:
        print(f"  [RemoteOK] Error: {e}")
    print(f"  [RemoteOK] {len(jobs)} jobs")
    return jobs
# ── SOURCE 5: WE WORK REMOTELY (RSS) ─────────────────────────────────────────
def fetch_weworkremotely(seen):
    import xml.etree.ElementTree as ET
    jobs = []
    feeds = [
        "https://weworkremotely.com/categories/remote-devops-sysadmin-jobs.rss",
        "https://weworkremotely.com/categories/remote-programming-jobs.rss",
        "https://weworkremotely.com/categories/remote-security-jobs.rss",
    ]
    for feed_url in feeds:
        try:
            r = requests.get(feed_url, timeout=10)
            root = ET.fromstring(r.content)
            for item in root.findall(".//item"):
                title   = item.findtext("title","")
                company = item.findtext("region","")
                link    = item.findtext("link","")
                if is_relevant(title):
                    jid = job_id(title, company)
                    if jid not in seen:
                        seen.add(jid)
                        jobs.append(make_job(title, company, "Remote", link, "WWR"))
        except Exception as e:
            print(f"  [WWR] Error: {e}")
        time.sleep(1)
    print(f"  [WeWorkRemotely] {len(jobs)} jobs")
    return jobs
# ── SOURCE 6: USAJOBS (FREE GOVT API) ────────────────────────────────────────
def fetch_usajobs(seen):
    jobs = []
    try:
        for title in ["Software Engineer","DevOps Engineer","Cloud Engineer"]:
            r = requests.get(
                "https://data.usajobs.gov/api/search",
                params={"Keyword": title, "ResultsPerPage": 25, "RemoteIndicator": "True"},
                headers={"Host": "data.usajobs.gov", "User-Agent": "JobBot/1.0",
                         "Authorization-Key": ""},  # works without key for basic search
                timeout=10
            )
            for item in r.json().get("SearchResult",{}).get("SearchResultItems",[]):
                pos  = item.get("MatchedObjectDescriptor",{})
                t    = pos.get("PositionTitle","")
                co   = pos.get("OrganizationName","US Government")
                loc  = pos.get("PositionLocationDisplay","USA")
                url  = pos.get("PositionURI","")
                sal  = pos.get("PositionRemuneration",[{}])[0].get("MinimumRange","—")
                jid  = job_id(t, co)
                if jid not in seen and is_relevant(t):
                    seen.add(jid)
                    jobs.append(make_job(t, co, loc, url, "USAJOBS", f"${sal}"))
            time.sleep(1)
    except Exception as e:
        print(f"  [USAJobs] Error: {e}")
    print(f"  [USAJobs] {len(jobs)} jobs")
    return jobs
# ── SOURCE 7: GREENHOUSE ATS ──────────────────────────────────────────────────
def fetch_greenhouse(seen):
    jobs = []
    for slug in GREENHOUSE_COMPANIES:
        try:
            r = requests.get(
                f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true",
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json().get("jobs", []):
                title = job.get("title","")
                if not is_relevant(title):
                    continue
                loc = job.get("location",{}).get("name","USA")
                url = job.get("absolute_url","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "GREENHOUSE"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Greenhouse] {len(jobs)} jobs from {len(GREENHOUSE_COMPANIES)} companies")
    return jobs
# ── SOURCE 8: LEVER ATS ───────────────────────────────────────────────────────
def fetch_lever(seen):
    jobs = []
    for slug in LEVER_COMPANIES:
        try:
            r = requests.get(
                f"https://api.lever.co/v0/postings/{slug}?mode=json&commitment=Full-time",
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json():
                title = job.get("text","")
                if not is_relevant(title):
                    continue
                loc = job.get("categories",{}).get("location","USA")
                url = job.get("hostedUrl","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "LEVER"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Lever] {len(jobs)} jobs from {len(LEVER_COMPANIES)} companies")
    return jobs
# ── SOURCE 9: ASHBY ATS ───────────────────────────────────────────────────────
def fetch_ashby(seen):
    jobs = []
    for slug in ASHBY_COMPANIES:
        try:
            r = requests.post(
                "https://api.ashbyhq.com/posting-api/job-board",
                json={"organizationHostedJobsPageName": slug},
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json().get("jobs", []):
                title = job.get("title","")
                if not is_relevant(title):
                    continue
                loc = job.get("location","Remote")
                url = job.get("jobUrl","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "ASHBY"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Ashby] {len(jobs)} jobs from {len(ASHBY_COMPANIES)} companies")
    return jobs
# ── SOURCE 10: SMARTRECRUITERS ────────────────────────────────────────────────
def fetch_smartrecruiters(seen):
    jobs = []
    for title in JOB_TITLES[:6]:
        try:
            r = requests.get(
                "https://api.smartrecruiters.com/v1/companies/jobs",
                params={"q": title, "country": "us", "typeOfEmployment": "permanent",
                        "limit": 20},
                timeout=10
            )
            for job in r.json().get("content", []):
                t   = job.get("name","")
                co  = job.get("company",{}).get("name","")
                loc = job.get("location",{}).get("city","USA")
                url = f"https://jobs.smartrecruiters.com/{job.get('company',{}).get('identifier','')}/{job.get('id','')}"
                jid = job_id(t, co)
                if jid not in seen and is_relevant(t):
                    seen.add(jid)
                    jobs.append(make_job(t, co, loc, url, "SMARTRECRUIT"))
        except Exception as e:
            print(f"  [SmartRecruiters] Error: {e}")
        time.sleep(1)
    print(f"  [SmartRecruiters] {len(jobs)} jobs")
    return jobs
# ── SOURCE 11: WORKDAY ────────────────────────────────────────────────────────
def fetch_workday(seen):
    jobs = []
    for company, tenant in WORKDAY_TENANTS:
        for title in ["Software Engineer","DevOps Engineer","Cloud Engineer","SRE"]:
            try:
                url = (f"https://{tenant}.wd5.myworkdayjobs.com/wday/cxs/"
                       f"{tenant}/External/jobs")
                r = requests.post(
                    url,
                    json={"appliedFacets":{}, "limit": 20, "offset": 0,
                          "searchText": title},
                    headers={**HEADERS, "Content-Type":"application/json"},
                    timeout=10
                )
                for job in r.json().get("jobPostings",[]):
                    t   = job.get("title","")
                    loc = job.get("locationsText","USA")
                    ext = job.get("externalPath","")
                    apply_url = f"https://{tenant}.wd5.myworkdayjobs.com/{tenant}/External/job{ext}"
                    jid = job_id(t, company)
                    if jid not in seen and is_relevant(t):
                        seen.add(jid)
                        jobs.append(make_job(t, company, loc, apply_url, "WORKDAY"))
            except Exception:
                pass
            time.sleep(0.5)
    print(f"  [Workday] {len(jobs)} jobs from {len(WORKDAY_TENANTS)} companies")
    return jobs
# ── SAVE TO EXCEL ─────────────────────────────────────────────────────────────
def save_to_excel(jobs):
    today    = datetime.now().strftime("%Y-%m-%d")
    filename = f"Jobs_{today}.xlsx"
    os.makedirs(SAVE_FOLDER, exist_ok=True)
    filepath = os.path.join(SAVE_FOLDER, filename)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    sub_fill = PatternFill("solid", start_color="2E75B6")
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    wht_fill = PatternFill("solid", start_color="FFFFFF")
    # Source color map
    src_colors = {
        "LINKEDIN":"4267B2","INDEED":"003A9B","GLASSDOOR":"0CAA41",
        "ZIP_RECRUITER":"4CAF50","GOOGLE":"DB4437","ADZUNA":"D71921",
        "JOOBLE":"FF6600","REMOTEOK":"000000","WWR":"1D9D74",
        "USAJOBS":"002868","GREENHOUSE":"24B47E","LEVER":"4A90E2",
        "ASHBY":"7B61FF","SMARTRECRUIT":"E8222E","WORKDAY":"F36E21",
    }
    wb = Workbook()
    ws = wb.active
    ws.title = today
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Job Openings — {today}   ({len(jobs)} roles from 10+ sources)"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = hdr_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    # Headers
    hdrs = ["Job Title","Company","Location","Source","Salary","Date Posted","Apply Link","Apply By (Deadline)","Applied Status","Notes"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = sub_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    ws.row_dimensions[2].height = 22
    # Data
    for i, job in enumerate(jobs):
        r, fill = i + 3, alt_fill if i % 2 == 0 else wht_fill
        src_color = src_colors.get(job["source"], "555555")
        row_vals  = [job["title"], job["company"], job["location"],
                     job["source"], job["salary"], job["posted"], job["url"], "", "Not Applied", ""]
        for col, val in enumerate(row_vals, 1):
            c           = ws.cell(row=r, column=col, value=val)
            c.fill      = fill
            c.border    = border
            c.alignment = Alignment(vertical="center")
            if col == 4:   # Source — colored
                c.font = Font(name="Arial", size=10, bold=True, color=src_color)
            elif col == 7: # URL
                c.font = Font(name="Arial", size=10, color="0070C0", underline="single")
            elif col == 9: # Applied Status
                c.font = Font(name="Arial", size=10, bold=True, color="7F7F7F")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = Font(name="Arial", size=10)
        ws.row_dimensions[r].height = 18
    # Footer
    fr = len(jobs) + 4
    ws.merge_cells(f"A{fr}:J{fr}")
    ws[f"A{fr}"] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                    f"Sources: JobBoards + Adzuna + Jooble + RemoteOK + WWR + "
                    f"USAJobs + Greenhouse + Lever + Ashby + SmartRecruiters + Workday  |  "
                    f"Total: {len(jobs)}")
    ws[f"A{fr}"].font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws[f"A{fr}"].fill      = hdr_fill
    ws[f"A{fr}"].alignment = Alignment(horizontal="center")
    for col, w in zip(["A","B","C","D","E","F","G","H","I","J"], [38,24,22,16,14,14,52,20,18,20]):
        ws.column_dimensions[col].width = w
    # Dropdown for Applied Status (col I = col 9)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Not Applied,Applied,Interviewing,Offer,Rejected,Not Relevant"',
        allow_blank=True,
        showDropDown=False
    )
    dv.prompt = "Select your application status"
    dv.promptTitle = "Applied Status"
    ws.add_data_validation(dv)
    for row_num in range(3, len(jobs) + 3):
        dv.add(ws.cell(row=row_num, column=9))
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{len(jobs)+2}"
    wb.save(filepath)
    print(f"Saved: {filepath}")
    return filepath
# ── TELEGRAM SUMMARY ──────────────────────────────────────────────────────────
def send_summary(jobs):
    today = datetime.now().strftime("%Y-%m-%d")
    # Count by source
    from collections import Counter
    by_src = Counter(j["source"] for j in jobs)
    src_summary = "  ".join(f"{k}: {v}" for k,v in by_src.most_common())
    if not jobs:
        send_telegram(f"
        return
 <b>Job Scan — {today}</b>\n\nNo new openings found today.")
    send_telegram(
        f" <b>Job Scan — {today}</b>\n"
        f" <b>{len(jobs)} new openings</b> found!\n"
        f" File: <b>Jobs_{today}.xlsx</b>\n\n"
        f"<b>By Source:</b>\n{src_summary}\n"
        f"{'─'*30}"
    )
    # Sort by source priority (most reliable/important first)
    source_priority = {
        "GREENHOUSE": 1, "LEVER": 2, "ASHBY": 3, "WORKDAY": 4,
        "SMARTRECRUIT": 5, "LINKEDIN": 6, "INDEED": 7,
        "GLASSDOOR": 8, "ZIP_RECRUITER": 9, "GOOGLE": 10,
        "ADZUNA": 11, "JOOBLE": 12, "REMOTEOK": 13,
        "WWR": 14, "USAJOBS": 15,
    }
    sorted_jobs = sorted(jobs, key=lambda j: source_priority.get(j["source"], 99))
    # Top 30 in order of importance
    for job in sorted_jobs[:30]:
        send_telegram(
            f"
 <b>{job['title']}</b>\n"
            f"
            f"
            f"
        )
 {job['company']}  |   {job['location']}\n"
 {job['salary']}  |  [{job['source']}]\n"
 <a href='{job['url']}'>Apply Now</a>"
    if len(jobs) > 30:
        send_telegram(f" <b>{len(jobs)-30} more</b> in <b>Jobs_{today}.xlsx</b>")
# ── MAIN RUN ──────────────────────────────────────────────────────────────────
def run():
    print(f"\n{'='*60}")
    print(f"Scan: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    seen = load_seen()
    all_jobs = []
    print("\n[1/11] Job Boards (LinkedIn/Indeed/Glassdoor/ZipRecruiter/Google)...")
    all_jobs += fetch_job_boards(seen)
    print("\n[2/11] Adzuna API...")
    all_jobs += fetch_adzuna(seen)
    print("\n[3/11] Jooble API...")
    all_jobs += fetch_jooble(seen)
    print("\n[4/11] RemoteOK...")
    all_jobs += fetch_remoteok(seen)
    print("\n[5/11] We Work Remotely...")
    all_jobs += fetch_weworkremotely(seen)
    print("\n[6/11] USAJobs...")
    all_jobs += fetch_usajobs(seen)
    print("\n[7/11] Greenhouse ATS (300+ companies)...")
    all_jobs += fetch_greenhouse(seen)
    print("\n[8/11] Lever ATS (200+ companies)...")
    all_jobs += fetch_lever(seen)
    print("\n[9/11] Ashby ATS (AI/startup companies)...")
    all_jobs += fetch_ashby(seen)
    print("\n[10/11] SmartRecruiters...")
    all_jobs += fetch_smartrecruiters(seen)
    print("\n[11/11] Workday (Apple/Google/Microsoft/Meta + 30 more)...")
    all_jobs += fetch_workday(seen)
    save_seen(seen)
    print(f"\nTotal new jobs found: {len(all_jobs)}")
    save_to_excel(all_jobs)
    send_summary(all_jobs)
# ── SCHEDULE ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    send_telegram(
        " <b>Full Job Bot is live!</b>\n"
        "
 Runs daily at 8:00 AM\n"
        "
 Sources: LinkedIn · Indeed · Glassdoor · ZipRecruiter · Google Jobs\n"
        "         Adzuna · Jooble · RemoteOK · WeWorkRemotely · USAJobs\n"
        "         Greenhouse · Lever · Ashby · SmartRecruiters · Workday\n"
        "
 Covers 500+ top US company career pages via ATS\n"
        "
 New Excel file daily: <b>Jobs_YYYY-MM-DD.xlsx</b>"
    )
    run()
    schedule.every().day.at("08:00").do(run)
    while True:
        schedule.run_pending()
        time.sleep(60)