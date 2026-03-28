"""
====================================================================
FULL JOB SEARCH BOT
====================================================================
Sources:
  1. Job Boards    — LinkedIn, Indeed, Glassdoor, ZipRecruiter, Google Jobs
  2. Free APIs     — Adzuna, Jooble, RemoteOK, USAJobs, WeWorkRemotely
  3. ATS Platforms — Greenhouse, Lever, Ashby, SmartRecruiters, Workday
                     (covers 500+ top US companies in one go)

Output:
  - New Excel file every day: Daily_Jobs/Jobs_YYYY-MM-DD.xlsx
  - Telegram alert with top openings + apply links

Setup:
  pip install python-jobspy openpyxl requests schedule pandas
  → Set TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID below
  → Optional: Set ADZUNA_APP_ID, ADZUNA_APP_KEY (free at developer.adzuna.com)
  → Optional: Set JOOBLE_API_KEY (free at jooble.org/api)
  → Run: python job_bot.py
====================================================================
"""

import os, json, time, hashlib, requests, schedule
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    from jobspy import scrape_jobs
    JOBSPY_OK = True
except ImportError:
    JOBSPY_OK = False
    print("Run: pip install python-jobspy")

# ── CONFIG ────────────────────────────────────────────────────────────────────
TELEGRAM_BOT_TOKEN = "8621053380:AAHGJhcJPARAnHfekKvHA7fQ2SE2CbpKnuo"
TELEGRAM_CHAT_ID   = "612269695"
SERPAPI_KEY        = "0dc3cd2139fce65ad7f82ccc12105acc89ff2f09a42518c46d73cfe57ca7b81b"
ADZUNA_APP_KEY     = os.getenv("ADZUNA_APP_KEY",      "")
JOOBLE_API_KEY     = os.getenv("JOOBLE_API_KEY",      "")   # free @ jooble.org/api
SAVE_FOLDER        = "Daily_Jobs"
SEEN_FILE          = "seen_jobs.json"

JOB_TITLES = [
    "Software Engineer", "Senior Software Engineer", "Staff Software Engineer",
    "DevOps Engineer", "Senior DevOps Engineer", "Platform Engineer",
    "Cloud Engineer", "Cloud Architect", "Infrastructure Engineer",
    "Site Reliability Engineer", "SRE", "Production Engineer",
    "Security Engineer", "DevSecOps Engineer", "MLOps Engineer",
    "Backend Engineer", "Full Stack Engineer", "Systems Engineer",
    "Kubernetes Engineer", "Cloud Security Engineer", "DataOps Engineer",
    "Software Development Engineer", "Application Engineer",
]

# ── 500+ COMPANIES via ATS ────────────────────────────────────────────────────
# Greenhouse: just the company slug → api.greenhouse.io/v1/boards/{slug}/jobs
GREENHOUSE_COMPANIES = [
    # Big Tech & Cloud
    "airbnb","coinbase","stripe","dropbox","square","lyft","doordash",
    "robinhood","brex","plaid","chime","affirm","marqeta","ripple",
    "databricks","snowflake","elastic","mongodb","twilio","okta",
    "cloudflare","crowdstrike","datadog","splunk","zscaler","netskope",
    "hashicorp","circleci","docker","gitlab","github","atlassian",
    "hubspot","zendesk","intercom","notion","figma","canva","asana",
    "airtable","clickup","linear","vercel","netlify","heroku",
    # AI / ML
    "anthropic","scale","cohere","huggingface","weights-biases",
    "openai","perplexity","mistral","together-ai","replicate",
    # Fintech
    "chime","sofi","nerdwallet","betterment","wealthfront","robinhood",
    "carta","rippling","gusto","adyen","payoneer","wise",
    # Health Tech
    "tempus","guardant","color-genomics","hims","ro","noom",
    "headspace","calm","cerebral","teladoc",
    # E-commerce / Consumer
    "etsy","wayfair","chewy","poshmark","offerup","mercari",
    "instacart","gopuff","getaround","turo","vacasa",
    # Enterprise SaaS
    "servicenow","workday","veeva","medallia","qualtrics","sprinklr",
    "amplitude","mixpanel","segment","mparticle","heap",
    "braze","klaviyo","sendgrid","mailchimp","iterable",
    # Infrastructure / DevOps
    "fastly","fly-io","render","railway","supabase","planetscale",
    "cockroachdb","yugabyte","citus","timescale","questdb",
    # Security
    "pagerduty","lacework","orca","wiz","snyk","sonatype",
    "sysdig","aqua-security","anchore","chainguard",
    # Gaming / Media
    "epicgames","roblox","unity","niantic","discord","twitch",
    # Logistics / Mobility
    "flexport","project44","motive","samsara","axon","rivian",
    "zoox","waymo","aurora","argo-ai",
    # IT Services
    "epam","thoughtworks","slalom","publicissapient",
    # Added from referrals + screenshots
    "adobe","intuit","veeva","paloaltonetworks","broadcom",
    "techmahindra","virtusa","servicenow","workday","hubspot",
    "okta","mongodb","twilio","zendesk","freshworks",
    "figma","notion","asana","linear","canva",
    "samsara","axon","rivian","flexport","project44",
    "wiz","lacework","orca","snyk","sonatype",
    "grafana","influxdata","honeycomb","lightstep","newrelic",
    "dbtlabs","fivetran","airbyte","alation","collibra",
]

# Lever: jobs.lever.co/{slug}?lever-source=jobspage
LEVER_COMPANIES = [
    # Big Tech & SaaS
    "netflix","lyft","atlassian","reddit","pinterest","quora",
    "shopify","squarespace","wix","bigcommerce","magento",
    "zendesk","freshworks","pipedrive","copper","close",
    "mixmax","outreach","salesloft","gong","chorus",
    # Cloud / Infra
    "hashicorp","vault","consul","terraform","nomad",
    "grafana","prometheus","influxdata","victoria-metrics",
    "dynatrace","newrelic","appdynamics","honeycomb","lightstep",
    # AI / Data
    "dbtlabs","fivetran","airbyte","stitch","matillion",
    "alation","atlan","collibra","informatica","talend",
    # Security
    "1password","dashlane","lastpass","bitwarden","keeper",
    "sentinelone","cylance","malwarebytes","bitdefender",
    # Fintech
    "navan","ramp","divvy","airbase","expensify","concur",
    "tipalti","bill","melio","paylocity","paycom",
    # Health
    "headway","alma","sondermind","talkspace","betterhelp",
    "sword-health","hinge-health","kaia-health","woebot",
    # E-comm / Marketplace
    "faire","angi","thumbtack","houzz","build","buildzoom",
    # Edtech
    "duolingo","coursera","udemy","pluralsight","skillshare",
    "masterclass","brilliant","khan-academy","codecademy",
    # Added missing referral companies
    "palo-alto-networks","veeva-systems","intuit","adobe",
    "broadcom","tech-mahindra","virtusa","vmware",
    "qualcomm","amd","intel","cisco","juniper",
]

# Ashby: jobs.ashbyhq.com/{slug}
ASHBY_COMPANIES = [
    "openai","perplexity","mistral","together","anyscale",
    "modal","modal-labs","fly","railway","render",
    "supabase","neon","planetscale","turso","xata",
    "resend","loops","postmark","sendgrid",
    "linear","height","plane","basecamp",
    "vercel","netlify","cloudflare-workers","deno",
    "cursor","codeium","tabnine","sourcegraph",
]

# SmartRecruiters public API
SMARTRECRUITERS_COMPANIES = [
    "Visa","Mastercard","PayPal","Intuit","Adobe","Salesforce",
    "Oracle","SAP","IBM","Cisco","Intel","AMD","Qualcomm","Broadcom",
    "Dell","HP","Lenovo","Logitech","Corsair",
    "Deloitte","PwC","EY","KPMG","Accenture","McKinsey","BCG",
    "JPMorgan","Goldman","Morgan Stanley","Citi","BankOfAmerica","Wells Fargo",
    "Amazon","Microsoft","Apple","Google","Meta","Nvidia","Tesla",
    "Walmart","Target","Costco","Kroger",
    "UnitedHealth","CVS","Cigna","Humana","Anthem",
    "Boeing","Lockheed","Raytheon","Northrop","GeneralDynamics",
]

# Workday tenant URLs (pattern: {tenant}.wd5.myworkdayjobs.com)
WORKDAY_TENANTS = [
    ("Apple",           "apple"),
    ("Google",          "google"),
    ("Microsoft",       "microsoft"),
    ("Meta",            "meta"),
    ("Amazon",          "amazon"),
    ("Nvidia",          "nvidia"),
    ("Tesla",           "tesla"),
    ("Salesforce",      "salesforce"),
    ("Adobe",           "adobe"),
    ("Intuit",          "intuit"),
    ("Workday",         "workday"),
    ("ServiceNow",      "servicenow"),
    ("Palo Alto",       "paloaltonetworks"),
    ("Crowdstrike",     "crowdstrike"),
    ("Zscaler",         "zscaler"),
    ("Snowflake",       "snowflake"),
    ("Databricks",      "databricks"),
    ("Splunk",          "splunk"),
    ("Elastic",         "elastic"),
    ("MongoDB",         "mongodb"),
    ("HubSpot",         "hubspot"),
    ("Okta",            "okta"),
    ("Twilio",          "twilio"),
    ("Zendesk",         "zendesk"),
    ("Veeva",           "veeva"),
    ("Informatica",     "informatica"),
    ("Cognizant",       "cognizant"),
    ("Infosys",         "infosys"),
    ("TCS",             "tcs"),
    ("Wipro",           "wipro"),
    ("HCL",             "hcl"),
    ("Capgemini",       "capgemini"),
    ("EPAM",            "epam"),
]

# ── HELPERS ───────────────────────────────────────────────────────────────────
HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}

def send_file(filepath):
    """Send Excel file directly to Telegram chat"""
    try:
        with open(filepath, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
                data={"chat_id": TELEGRAM_CHAT_ID,
                      "caption": f"📂 Daily Jobs Sheet — {datetime.now().strftime('%Y-%m-%d')}"},
                files={"document": f},
                timeout=60
            )
        print(f"File sent to Telegram: {filepath}")
    except Exception as e:
        print(f"  File send error: {e}")

# ── SAI SIVANI'S RESUME PROFILE ──────────────────────────────────────────────
RESUME_SKILLS = {
    # Cloud (weight 3 — strongest area)
    "aws": 3, "ec2": 3, "eks": 3, "s3": 3, "lambda": 3, "iam": 3,
    "vpc": 3, "cloudwatch": 3, "rds": 3, "route53": 3, "cloudformation": 3,
    "codepipeline": 3, "auto scaling": 3,
    "azure": 3, "aks": 3, "key vault": 3, "azure devops": 3,
    # Containers & Orchestration (weight 3)
    "kubernetes": 3, "k8s": 3, "docker": 3, "helm": 3, "argocd": 3,
    "gitops": 3, "eks": 3, "aks": 3,
    # CI/CD (weight 3)
    "jenkins": 3, "github actions": 3, "gitlab ci": 3, "gitlab": 3,
    "ci/cd": 3, "cicd": 3, "pipeline": 2,
    # IaC (weight 3)
    "terraform": 3, "ansible": 3, "infrastructure as code": 3, "iac": 3,
    # Observability (weight 2)
    "prometheus": 2, "grafana": 2, "datadog": 2, "elk": 2, "splunk": 2,
    "cloudwatch": 2, "opentelemetry": 2, "monitoring": 2, "observability": 2,
    # Security / DevSecOps (weight 2)
    "trivy": 2, "snyk": 2, "sonarqube": 2, "devsecops": 2,
    "rbac": 2, "secrets management": 2, "vault": 2, "security": 2,
    # Languages (weight 2)
    "python": 2, "bash": 2, "boto3": 2, "powershell": 2,
    # Practices (weight 1)
    "agile": 1, "scrum": 1, "on-call": 1, "incident management": 1,
    "cost optimization": 1, "sre": 2, "site reliability": 2,
    # Certs (weight 1)
    "aws certified": 1, "solutions architect": 1,
}

DEVOPS_CLOUD_ROLES = [
    "devops", "sre", "site reliability", "platform engineer",
    "cloud engineer", "cloud architect", "infrastructure engineer",
    "kubernetes engineer", "devsecops", "mlops", "dataops",
    "reliability engineer", "build engineer", "release engineer",
]

def is_devops_cloud_role(title):
    t = title.lower()
    return any(r in t for r in DEVOPS_CLOUD_ROLES)

def score_job(title, company, description=""):
    """Score job 1-10 based on match with Sai's resume. Only for DevOps/Cloud roles."""
    if not is_devops_cloud_role(title):
        return ""  # blank for non DevOps/Cloud roles

    text = f"{title} {company} {description}".lower()
    total_score = 0
    max_possible = 0
    matched = []

    for skill, weight in RESUME_SKILLS.items():
        max_possible += weight
        if skill in text:
            total_score += weight
            matched.append(skill)

    if max_possible == 0:
        return "5/10"

    raw = (total_score / max_possible) * 10
    score = max(1, min(10, round(raw)))
    return f"{score}/10"

# ── COMPANY TIERS ────────────────────────────────────────────────────────────
TIER_1_COMPANIES = {
    # Big Tech
    "google","meta","apple","microsoft","amazon","nvidia","netflix",
    # Top DevOps/Cloud
    "datadog","hashicorp","cloudflare","crowdstrike","snowflake","databricks",
    "elastic","splunk","zscaler","netskope","okta","mongodb","twilio",
    "pagerduty","grafana","circleci","docker","gitlab","github",
    # Top Fintech
    "stripe","coinbase","robinhood","brex","plaid","chime","affirm",
    # Top AI
    "anthropic","openai","scale","cohere",
    # Top SaaS
    "salesforce","hubspot","servicenow","workday","atlassian","zendesk",
    "notion","figma","canva","asana","linear",
}

TIER_2_COMPANIES = {
    "airbnb","lyft","doordash","instacart","pinterest","reddit","discord",
    "dropbox","box","zoom","slack","intercom","freshworks","zendesk",
    "fastly","fly","render","supabase","planetscale","cockroachdb",
    "rippling","gusto","carta","ramp","navan","divvy",
    "samsara","axon","rivian","flexport","project44",
    "cohesity","wasabi","cloudapp","dnsimple","fairwinds",
}

ROLE_PRIORITY = {
    # DevOps/Cloud entry-mid first
    "devops engineer": 1, "cloud engineer": 1,
    "site reliability engineer": 1, "sre": 1,
    "platform engineer": 1, "infrastructure engineer": 1,
    "kubernetes engineer": 1, "devsecops engineer": 1,
    "mlops engineer": 2, "dataops engineer": 2,
    "cloud architect": 2, "solutions architect": 2,
    # SWE roles after
    "software engineer": 3, "backend engineer": 3,
    "full stack engineer": 3, "systems engineer": 4,
    "security engineer": 4, "application engineer": 5,
    "data engineer": 5,
}

def get_company_tier(company):
    co = company.lower().strip()
    if any(t in co for t in TIER_1_COMPANIES):
        return 1
    elif any(t in co for t in TIER_2_COMPANIES):
        return 2
    return 3

def get_role_priority(title):
    t = title.lower()
    for role, priority in ROLE_PRIORITY.items():
        if role in t:
            return priority
    return 6

def get_exp_sort(title):
    """0-2 yrs = 0 (top), 0-5 yrs = 1 (bottom)"""
    t = title.lower()
    if any(x in t for x in ["junior","jr.","jr ","entry","associate"," i-"," i "]):
        return 0
    elif any(x in t for x in ["mid","ii ","level 2","l2","intermediate"]):
        return 1
    return 2  # general/unspecified — after entry, before pure 0-5

def get_sort_score(job):
    """Lower score = higher priority"""
    tier      = get_company_tier(job["company"])
    role_pri  = get_role_priority(job["title"])
    exp_sort  = get_exp_sort(job["title"])
    ms        = score_job(job["title"], job["company"])
    match_inv = 10 - int(ms.split("/")[0]) if ms and "/" in str(ms) else 5
    has_sal   = 0 if job["salary"] != "—" else 1
    # Sort: Role → Exp level → Match score → Company tier
    return (role_pri, exp_sort, match_inv, tier, has_sal)

def is_usa_or_remote(location):
    """Strictly keep only USA and Remote jobs"""
    loc = location.lower().strip()

    # Explicit remote — always keep
    if any(x in loc for x in ["remote", "work from home", "wfh", "anywhere"]):
        return True

    # Must contain usa indicators
    usa_indicators = [
        "usa", "united states", "u.s.", "us ", ", us",
        "alabama","alaska","arizona","arkansas","california","colorado",
        "connecticut","delaware","florida","georgia","hawaii","idaho",
        "illinois","indiana","iowa","kansas","kentucky","louisiana",
        "maine","maryland","massachusetts","michigan","minnesota",
        "mississippi","missouri","montana","nebraska","nevada",
        "new hampshire","new jersey","new mexico","new york",
        "north carolina","north dakota","ohio","oklahoma","oregon",
        "pennsylvania","rhode island","south carolina","south dakota",
        "tennessee","texas","utah","vermont","virginia","washington",
        "west virginia","wisconsin","wyoming",
        " al"," ak"," az"," ar"," ca"," co"," ct"," de"," fl"," ga",
        " hi"," id"," il"," in"," ia"," ks"," ky"," la"," me"," md",
        " ma"," mi"," mn"," ms"," mo"," mt"," ne"," nv"," nh"," nj",
        " nm"," ny"," nc"," nd"," oh"," ok"," or"," pa"," ri"," sc",
        " sd"," tn"," tx"," ut"," vt"," va"," wa"," wv"," wi"," wy",
    ]
    if any(x in loc for x in usa_indicators):
        return True

    # Explicitly exclude non-USA
    exclude = [
        "canada","toronto","vancouver","montreal","ontario","british columbia",
        "uk","london","manchester","edinburgh","birmingham","england","scotland",
        "germany","berlin","munich","hamburg","frankfurt",
        "india","bangalore","mumbai","delhi","hyderabad","pune","chennai",
        "australia","sydney","melbourne","brisbane","perth",
        "brazil","sao paulo","rio","amsterdam","netherlands","paris","france",
        "spain","madrid","barcelona","italy","rome","milan",
        "poland","warsaw","portugal","lisbon","ireland","dublin",
        "singapore","hong kong","japan","tokyo","china","beijing",
        "mexico","colombia","argentina","chile","peru",
        "philippines","pakistan","bangladesh","nigeria","kenya",
    ]
    if any(c in loc for c in exclude):
        return False

    # If location is empty or just "remote" variants — keep
    if not loc or len(loc) < 3:
        return True

    return False

def is_entry_mid_level(title, description=""):
    """Strictly filter for 0-5 years / junior-mid level roles only"""
    title_lower = title.lower()

    # Strictly exclude all senior/lead/staff/principal/architect roles
    exclude = [
        "senior", "sr.", "sr ", " sr-", "staff ", "principal",
        "director", "vp ", "vice president", "head of",
        "distinguished", "fellow", "architect", "lead ",
        "manager", "management", " iv", " v ", "level 5",
        "level 6", "l5", "l6", "10+ years", "8+ years",
        "7+ years", "6+ years", "tech lead", "team lead",
    ]
    for word in exclude:
        if word in title_lower:
            return False

    return True

def get_exp_level(title):
    """Return experience level label for 0-5 yr roles"""
    title_lower = title.lower()
    if any(x in title_lower for x in ["junior", "jr.", "jr ", "entry", "associate", "i ", " i-"]):
        return "0-2 yrs"
    elif any(x in title_lower for x in ["mid", "ii ", "level 2", "l2", "intermediate"]):
        return "2-4 yrs"
    else:
        return "0-5 yrs"

def load_seen():
    if os.path.exists(SEEN_FILE):
        with open(SEEN_FILE) as f:
            return set(json.load(f))
    return set()

def save_seen(seen):
    with open(SEEN_FILE, "w") as f:
        json.dump(list(seen), f)

def job_id(title, company):
    return hashlib.md5(f"{title}{company}".lower().encode()).hexdigest()

def make_job(title, company, location, url, source, salary="—"):
    return {"title": title, "company": company, "location": location,
            "url": url, "source": source, "salary": salary,
            "posted": datetime.now().strftime("%Y-%m-%d")}

def is_good_job(title, location):
    """Check if job matches our criteria"""
    if not is_relevant(title):
        return False
    if not is_usa_or_remote(location):
        return False
    if not is_entry_mid_level(title):
        return False
    return True

def send_telegram(msg):
    try:
        requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            json={"chat_id": TELEGRAM_CHAT_ID, "text": msg,
                  "parse_mode": "HTML", "disable_web_page_preview": True},
            timeout=10
        )
        time.sleep(0.4)
    except Exception as e:
        print(f"  Telegram error: {e}")

def is_relevant(title):
    keywords = [
        "engineer","sre","devops","platform","cloud","infrastructure",
        "backend","fullstack","full stack","security","mlops","dataops",
        "reliability","kubernetes","architect","systems"
    ]
    return any(k in title.lower() for k in keywords)

# ── SOURCE 1: JOB BOARDS (JobSpy) ────────────────────────────────────────────
def fetch_job_boards(seen):
    if not JOBSPY_OK:
        return []
    import pandas as pd
    jobs, frames = [], []
    for title in JOB_TITLES:
        print(f"  [JobBoards] {title}")
        try:
            df = scrape_jobs(
                site_name=["linkedin","indeed","glassdoor","zip_recruiter","google"],
                search_term=title,
                location="United States",
                results_wanted=25,
                hours_old=24,
                is_remote=True,
                job_type="fulltime",
                country_indeed="USA",
            )
            if not df.empty:
                frames.append(df)
        except Exception as e:
            print(f"    Error: {e}")
        time.sleep(2)

    if not frames:
        return []

    for _, row in pd.concat(frames, ignore_index=True).iterrows():
        jid = job_id(str(row.get("title","")), str(row.get("company","")))
        if jid not in seen:
            seen.add(jid)
            jobs.append(make_job(
                str(row.get("title","")), str(row.get("company","")),
                str(row.get("location","Remote")),
                str(row.get("job_url", row.get("job_link",""))),
                str(row.get("site","")).upper(),
                str(row.get("min_amount","") or "—"),
            ))
    return jobs

# ── SOURCE 2: ADZUNA API ──────────────────────────────────────────────────────
def fetch_adzuna(seen):
    if not ADZUNA_APP_ID:
        print("  [Adzuna] Skipping — no API key (get free key at developer.adzuna.com)")
        return []
    jobs = []
    for title in JOB_TITLES[:8]:  # limit to avoid rate limits
        try:
            url = (f"https://api.adzuna.com/v1/api/jobs/us/search/1"
                   f"?app_id={ADZUNA_APP_ID}&app_key={ADZUNA_APP_KEY}"
                   f"&results_per_page=20&what={requests.utils.quote(title)}"
                   f"&where=remote&max_days_old=1&content-type=application/json")
            r = requests.get(url, timeout=10)
            for job in r.json().get("results", []):
                jid = job_id(job.get("title",""), job.get("company",{}).get("display_name",""))
                if jid not in seen and is_relevant(job.get("title","")):
                    seen.add(jid)
                    jobs.append(make_job(
                        job.get("title",""),
                        job.get("company",{}).get("display_name",""),
                        job.get("location",{}).get("display_name","Remote"),
                        job.get("redirect_url",""),
                        "ADZUNA",
                        f"${job.get('salary_min',0):,.0f}" if job.get("salary_min") else "—",
                    ))
        except Exception as e:
            print(f"  [Adzuna] Error: {e}")
        time.sleep(1)
    print(f"  [Adzuna] {len(jobs)} jobs")
    return jobs

# ── SOURCE 3: JOOBLE API ──────────────────────────────────────────────────────
def fetch_jooble(seen):
    if not JOOBLE_API_KEY:
        print("  [Jooble] Skipping — no API key (get free key at jooble.org/api)")
        return []
    jobs = []
    for title in JOB_TITLES[:8]:
        try:
            r = requests.post(
                f"https://jooble.org/api/{JOOBLE_API_KEY}",
                json={"keywords": title, "location": "remote", "page": "1"},
                timeout=10
            )
            for job in r.json().get("jobs", []):
                jid = job_id(job.get("title",""), job.get("company",""))
                if jid not in seen and is_relevant(job.get("title","")):
                    seen.add(jid)
                    jobs.append(make_job(
                        job.get("title",""), job.get("company",""),
                        job.get("location","Remote"),
                        job.get("link",""), "JOOBLE",
                        job.get("salary","—"),
                    ))
        except Exception as e:
            print(f"  [Jooble] Error: {e}")
        time.sleep(1)
    print(f"  [Jooble] {len(jobs)} jobs")
    return jobs

# ── SOURCE 4: REMOTEOK (FREE PUBLIC API) ─────────────────────────────────────
def fetch_remoteok(seen):
    jobs = []
    try:
        r = requests.get("https://remoteok.com/api", headers={"User-Agent": "JobBot/1.0"}, timeout=15)
        for job in r.json()[1:]:  # first item is metadata
            if not is_relevant(job.get("position","")):
                continue
            jid = job_id(job.get("position",""), job.get("company",""))
            if jid not in seen:
                seen.add(jid)
                salary = ""
                if job.get("salary_min"):
                    salary = f"${job['salary_min']:,} - ${job.get('salary_max',0):,}"
                jobs.append(make_job(
                    job.get("position",""), job.get("company",""),
                    "Remote", job.get("url",""), "REMOTEOK", salary or "—",
                ))
    except Exception as e:
        print(f"  [RemoteOK] Error: {e}")
    print(f"  [RemoteOK] {len(jobs)} jobs")
    return jobs

# ── SOURCE 5: WE WORK REMOTELY (RSS) ─────────────────────────────────────────
def fetch_weworkremotely(seen):
    import xml.etree.ElementTree as ET
    jobs = []
    feeds = [
        "https://weworkremotely.com/categories/remote-devops-sysadmin-jobs.rss",
        "https://weworkremotely.com/categories/remote-programming-jobs.rss",
        "https://weworkremotely.com/categories/remote-security-jobs.rss",
    ]
    for feed_url in feeds:
        try:
            r = requests.get(feed_url, timeout=10)
            root = ET.fromstring(r.content)
            for item in root.findall(".//item"):
                title   = item.findtext("title","")
                company = item.findtext("region","")
                link    = item.findtext("link","")
                if is_relevant(title):
                    jid = job_id(title, company)
                    if jid not in seen:
                        seen.add(jid)
                        jobs.append(make_job(title, company, "Remote", link, "WWR"))
        except Exception as e:
            print(f"  [WWR] Error: {e}")
        time.sleep(1)
    print(f"  [WeWorkRemotely] {len(jobs)} jobs")
    return jobs

# ── SOURCE 6: USAJOBS (FREE GOVT API) ────────────────────────────────────────
def fetch_usajobs(seen):
    jobs = []
    try:
        for title in ["Software Engineer","DevOps Engineer","Cloud Engineer"]:
            r = requests.get(
                "https://data.usajobs.gov/api/search",
                params={"Keyword": title, "ResultsPerPage": 25, "RemoteIndicator": "True"},
                headers={"Host": "data.usajobs.gov", "User-Agent": "JobBot/1.0",
                         "Authorization-Key": ""},  # works without key for basic search
                timeout=10
            )
            for item in r.json().get("SearchResult",{}).get("SearchResultItems",[]):
                pos  = item.get("MatchedObjectDescriptor",{})
                t    = pos.get("PositionTitle","")
                co   = pos.get("OrganizationName","US Government")
                loc  = pos.get("PositionLocationDisplay","USA")
                url  = pos.get("PositionURI","")
                sal  = pos.get("PositionRemuneration",[{}])[0].get("MinimumRange","—")
                jid  = job_id(t, co)
                if jid not in seen and is_relevant(t):
                    seen.add(jid)
                    jobs.append(make_job(t, co, loc, url, "USAJOBS", f"${sal}"))
            time.sleep(1)
    except Exception as e:
        print(f"  [USAJobs] Error: {e}")
    print(f"  [USAJobs] {len(jobs)} jobs")
    return jobs

# ── SOURCE 7: GREENHOUSE ATS ──────────────────────────────────────────────────
def fetch_greenhouse(seen):
    jobs = []
    for slug in GREENHOUSE_COMPANIES:
        try:
            r = requests.get(
                f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true",
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json().get("jobs", []):
                title = job.get("title","")
                if not is_relevant(title):
                    continue
                loc = job.get("location",{}).get("name","USA")
                url = job.get("absolute_url","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "GREENHOUSE"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Greenhouse] {len(jobs)} jobs from {len(GREENHOUSE_COMPANIES)} companies")
    return jobs

# ── SOURCE 8: LEVER ATS ───────────────────────────────────────────────────────
def fetch_lever(seen):
    jobs = []
    for slug in LEVER_COMPANIES:
        try:
            r = requests.get(
                f"https://api.lever.co/v0/postings/{slug}?mode=json&commitment=Full-time",
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json():
                title = job.get("text","")
                if not is_relevant(title):
                    continue
                loc = job.get("categories",{}).get("location","USA")
                url = job.get("hostedUrl","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "LEVER"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Lever] {len(jobs)} jobs from {len(LEVER_COMPANIES)} companies")
    return jobs

# ── SOURCE 9: ASHBY ATS ───────────────────────────────────────────────────────
def fetch_ashby(seen):
    jobs = []
    for slug in ASHBY_COMPANIES:
        try:
            r = requests.post(
                "https://api.ashbyhq.com/posting-api/job-board",
                json={"organizationHostedJobsPageName": slug},
                timeout=10
            )
            if r.status_code != 200:
                continue
            for job in r.json().get("jobs", []):
                title = job.get("title","")
                if not is_relevant(title):
                    continue
                loc = job.get("location","Remote")
                url = job.get("jobUrl","")
                jid = job_id(title, slug)
                if jid not in seen:
                    seen.add(jid)
                    jobs.append(make_job(title, slug.title(), loc, url, "ASHBY"))
        except Exception as e:
            pass
        time.sleep(0.3)
    print(f"  [Ashby] {len(jobs)} jobs from {len(ASHBY_COMPANIES)} companies")
    return jobs

# ── SOURCE 10: SMARTRECRUITERS ────────────────────────────────────────────────
def fetch_smartrecruiters(seen):
    jobs = []
    for title in JOB_TITLES[:6]:
        try:
            r = requests.get(
                "https://api.smartrecruiters.com/v1/companies/jobs",
                params={"q": title, "country": "us", "typeOfEmployment": "permanent",
                        "limit": 20},
                timeout=10
            )
            for job in r.json().get("content", []):
                t   = job.get("name","")
                co  = job.get("company",{}).get("name","")
                loc = job.get("location",{}).get("city","USA")
                url = f"https://jobs.smartrecruiters.com/{job.get('company',{}).get('identifier','')}/{job.get('id','')}"
                jid = job_id(t, co)
                if jid not in seen and is_relevant(t):
                    seen.add(jid)
                    jobs.append(make_job(t, co, loc, url, "SMARTRECRUIT"))
        except Exception as e:
            print(f"  [SmartRecruiters] Error: {e}")
        time.sleep(1)
    print(f"  [SmartRecruiters] {len(jobs)} jobs")
    return jobs

# ── SOURCE 11: WORKDAY ────────────────────────────────────────────────────────
def fetch_workday(seen):
    jobs = []
    for company, tenant in WORKDAY_TENANTS:
        for title in ["Software Engineer","DevOps Engineer","Cloud Engineer","SRE"]:
            try:
                url = (f"https://{tenant}.wd5.myworkdayjobs.com/wday/cxs/"
                       f"{tenant}/External/jobs")
                r = requests.post(
                    url,
                    json={"appliedFacets":{}, "limit": 20, "offset": 0,
                          "searchText": title},
                    headers={**HEADERS, "Content-Type":"application/json"},
                    timeout=10
                )
                for job in r.json().get("jobPostings",[]):
                    t   = job.get("title","")
                    loc = job.get("locationsText","USA")
                    ext = job.get("externalPath","")
                    apply_url = f"https://{tenant}.wd5.myworkdayjobs.com/{tenant}/External/job{ext}"
                    jid = job_id(t, company)
                    if jid not in seen and is_relevant(t):
                        seen.add(jid)
                        jobs.append(make_job(t, company, loc, apply_url, "WORKDAY"))
            except Exception:
                pass
            time.sleep(0.5)
    print(f"  [Workday] {len(jobs)} jobs from {len(WORKDAY_TENANTS)} companies")
    return jobs

# ── SAVE TO EXCEL ─────────────────────────────────────────────────────────────
def save_to_excel(jobs):
    today    = datetime.now().strftime("%Y-%m-%d")
    filename = f"Jobs_{today}.xlsx"
    os.makedirs(SAVE_FOLDER, exist_ok=True)
    filepath = os.path.join(SAVE_FOLDER, filename)

    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    sub_fill = PatternFill("solid", start_color="2E75B6")
    alt_fill = PatternFill("solid", start_color="EBF3FB")
    wht_fill = PatternFill("solid", start_color="FFFFFF")

    # Source color map
    src_colors = {
        "LINKEDIN":"4267B2","INDEED":"003A9B","GLASSDOOR":"0CAA41",
        "ZIP_RECRUITER":"4CAF50","GOOGLE":"DB4437","ADZUNA":"D71921",
        "JOOBLE":"FF6600","REMOTEOK":"000000","WWR":"1D9D74",
        "USAJOBS":"002868","GREENHOUSE":"24B47E","LEVER":"4A90E2",
        "ASHBY":"7B61FF","SMARTRECRUIT":"E8222E","WORKDAY":"F36E21",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = today

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Job Openings — {today}   ({len(jobs)} roles found)"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = hdr_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Headers
    hdrs = ["Job Title","Company","Tier","H-1B","Location","Source","Salary","Exp Level","Match Score","Date Posted","Apply Link","Apply By (Deadline)","Applied Status","Notes"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill      = sub_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
    ws.row_dimensions[2].height = 22

    # Data
    for i, job in enumerate(jobs):
        r, fill = i + 3, alt_fill if i % 2 == 0 else wht_fill
        src_color = src_colors.get(job["source"], "555555")
        exp_level   = get_exp_level(job["title"])
        tier        = ["⭐⭐⭐ Top Tier","⭐⭐ Mid Tier","⭐ Other"][get_company_tier(job["company"])-1]
        match_score = score_job(job["title"], job["company"])
        h1b_cos     = {'hubspot', 'cisco', 'airbnb', 'splunk', 'microsoft', 'carta', 'tcs', 'rivian', 'databricks', 'wipro', 'broadcom', 'sofi', 'intuit', 'zscaler', 'samsara', 'spacex', 'hcl', 'netskope', 'epam', 'chewy', 'toast', 'tesla', 'openai', 'qualcomm', 'servicenow', 'guardant', 'nvidia', 'oracle', 'ripple', 'zoom', 'clearwater', 'mongodb', 'slalom', 'amazon', 'capgemini', 'coinbase', 'scale', 'plaid', 'lyft', 'wayfair', 'axon', 'cloudflare', 'cohesity', 'adobe', 'zipline', 'chime', 'affirm', 'google', 'dataminr', 'stripe', 'tempus', 'flexport', 'cognizant', 'servicetitan', 'palo alto', 'datadog', 'verkada', 'ginkgo', 'elastic', 'rippling', 'anthropic', 'brex', 'okta', 'workday', 'infosys', 'palantir', 'meta', 'ltimindtree', 'robinhood', 'snowflake', 'tiktok', 'doordash', 'twilio', 'instacart', 'marqeta', 'hashicorp', 'crowdstrike', 'salesforce', 'apple'}
        h1b         = "✅ Yes" if any(h in job["company"].lower() for h in h1b_cos) else "—"
        row_vals    = [job["title"], job["company"], tier, h1b, job["location"],
                       job["source"], job["salary"], exp_level, match_score,
                       job["posted"], job["url"], "", "Not Applied", ""]
        for col, val in enumerate(row_vals, 1):
            c           = ws.cell(row=r, column=col, value=val)
            c.fill      = fill
            c.border    = border
            c.alignment = Alignment(vertical="center")
            if col == 3:   # Tier
                tier_colors = {"⭐⭐⭐ Top Tier":"7B2C2C","⭐⭐ Mid Tier":"375623","⭐ Other":"1F4E79"}
                c.font = Font(name="Arial", size=10, bold=True, color=tier_colors.get(val,"000000"))
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 4: # H-1B
                c.font = Font(name="Arial", size=10, bold=True,
                              color="375623" if val == "✅ Yes" else "AAAAAA")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 6:   # Source — colored
                c.font = Font(name="Arial", size=10, bold=True, color=src_color)
            elif col == 8: # Exp Level
                c.font = Font(name="Arial", size=10, bold=True, color="375623")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 9: # Match Score
                score_num = int(val.split("/")[0]) if val and "/" in str(val) else 0
                score_color = "375623" if score_num >= 7 else ("FF6600" if score_num >= 4 else "7F7F7F")
                c.font = Font(name="Arial", size=11, bold=True, color=score_color)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 11: # URL
                c.font = Font(name="Arial", size=10, color="0070C0", underline="single")
            elif col == 13: # Applied Status
                c.font = Font(name="Arial", size=10, bold=True, color="7F7F7F")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = Font(name="Arial", size=10)
        ws.row_dimensions[r].height = 18

    # Footer
    fr = len(jobs) + 4
    ws.merge_cells(f"A{fr}:N{fr}")
    ws[f"A{fr}"] = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                    f"Sources: JobBoards + Adzuna + Jooble + RemoteOK + WWR + "
                    f"USAJobs + Greenhouse + Lever + Ashby + SmartRecruiters + Workday  |  "
                    f"Total: {len(jobs)}")
    ws[f"A{fr}"].font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    ws[f"A{fr}"].fill      = hdr_fill
    ws[f"A{fr}"].alignment = Alignment(horizontal="center")

    for col, w in zip(["A","B","C","D","E","F","G","H","I","J","K","L","M","N"], [36,22,14,10,20,14,14,12,12,14,50,20,18,20]):
        ws.column_dimensions[col].width = w

    # Dropdown for Applied Status (col I = col 9)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Not Applied,Applied,Interviewing,Offer,Rejected,Not Relevant"',
        allow_blank=True,
        showDropDown=False
    )
    dv.prompt = "Select your application status"
    dv.promptTitle = "Applied Status"
    ws.add_data_validation(dv)
    for row_num in range(3, len(jobs) + 3):
        dv.add(ws.cell(row=row_num, column=13))

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{len(jobs)+2}"

    wb.save(filepath)
    print(f"Saved: {filepath}")
    return filepath

# ── TELEGRAM SUMMARY ──────────────────────────────────────────────────────────
def send_summary(jobs):
    """Send a clean summary to Telegram — Excel file will follow"""
    from collections import Counter
    today = datetime.now().strftime("%Y-%m-%d")

    if not jobs:
        send_telegram(f"📋 <b>Job Scan — {today}</b>\n\nNo new openings found today.")
        return

    # Count DevOps/Cloud matches
    devops_kw = ["devops","sre","site reliability","platform engineer","cloud engineer",
                 "infrastructure engineer","kubernetes","devsecops","mlops","dataops"]
    devops_count = sum(1 for j in jobs if any(k in j["title"].lower() for k in devops_kw))

    # Count by source
    by_src = Counter(j["source"] for j in jobs)
    src_summary = "\n".join(f"  • {k}: {v}" for k, v in by_src.most_common())

    send_telegram(
        f"📋 <b>Job Scan Complete — {today}</b>\n\n"
        f"✅ <b>{len(jobs)}</b> total new openings found\n"
        f"🎯 <b>{devops_count}</b> DevOps/Cloud roles\n\n"
        f"<b>By Source:</b>\n{src_summary}\n\n"
        f"📂 Your Excel file is coming right up!"
    )

# ── SOURCE 12: SERPAPI (ZipRecruiter + Glassdoor + Monster + Google Jobs + Workday companies) ──
def fetch_serpapi(seen):
    if not SERPAPI_KEY:
        print("  [SerpAPI] Skipping — no API key")
        return []
    jobs = []

    # ⚡ Only 3 searches per day to preserve SerpAPI free tier (100/month)
    # 1 broad DevOps/Cloud search + 1 SWE search + 1 top company search
    all_queries = [
        "DevOps Cloud SRE Platform Engineer remote USA",
        "Software Backend Infrastructure Engineer remote USA",
        "Apple Google Microsoft Meta Adobe Intuit DevOps Cloud Engineer remote",
    ]

    for query in all_queries:
        try:
            r = requests.get(
                "https://serpapi.com/search",
                params={
                    "engine":   "google_jobs",
                    "q":        query,
                    "location": "United States",
                    "hl":       "en",
                    "gl":       "us",
                    "chips":    "date_posted:today,employment_type:FULLTIME",
                    "ltype":    "1",  # remote jobs only
                    "api_key":  SERPAPI_KEY,
                },
                timeout=15
            )
            data = r.json()
            for job in data.get("jobs_results", []):
                t   = job.get("title", "")
                co  = job.get("company_name", "")
                loc = job.get("location", "Remote, USA")

                # Get best apply link
                url = ""
                for opt in job.get("apply_options", []):
                    url = opt.get("link", "")
                    break
                if not url:
                    url = f"https://www.google.com/search?q={requests.utils.quote(t+' '+co+' jobs')}"

                # Get salary if available
                salary = "—"
                for h in job.get("job_highlights", []):
                    for item in h.get("items", []):
                        if "$" in item or "salary" in item.lower():
                            salary = item[:50]
                            break

                jid = job_id(t, co)
                if jid not in seen and is_good_job(t, loc) and is_usa_or_remote(loc):
                    seen.add(jid)
                    jobs.append(make_job(t, co, loc, url, "SERPAPI", salary))
        except Exception as e:
            print(f"  [SerpAPI] Error on '{query}': {e}")
        time.sleep(1)

    print(f"  [SerpAPI] {len(jobs)} jobs (ZipRecruiter + Glassdoor + Monster + Google + Workday companies)")
    return jobs

# ── MAIN RUN ──────────────────────────────────────────────────────────────────
def run():
    print(f"\n{'='*60}")
    print(f"Scan: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    seen = load_seen()
    all_jobs = []

    print("\n[1/11] Job Boards (LinkedIn/Indeed/Glassdoor/ZipRecruiter/Google)...")
    all_jobs += fetch_job_boards(seen)

    print("\n[2/11] Adzuna API...")
    all_jobs += fetch_adzuna(seen)

    print("\n[3/11] Jooble API...")
    all_jobs += fetch_jooble(seen)

    print("\n[4/11] RemoteOK...")
    all_jobs += fetch_remoteok(seen)

    print("\n[5/11] We Work Remotely...")
    all_jobs += fetch_weworkremotely(seen)

    print("\n[6/11] USAJobs...")
    all_jobs += fetch_usajobs(seen)

    print("\n[7/11] Greenhouse ATS (300+ companies)...")
    all_jobs += fetch_greenhouse(seen)

    print("\n[8/11] Lever ATS (200+ companies)...")
    all_jobs += fetch_lever(seen)

    print("\n[9/11] Ashby ATS (AI/startup companies)...")
    all_jobs += fetch_ashby(seen)

    print("\n[10/11] SmartRecruiters...")
    all_jobs += fetch_smartrecruiters(seen)

    print("\n[11/11] Workday (Apple/Google/Microsoft/Meta + 30 more)...")
    all_jobs += fetch_workday(seen)

    print("\n[12/12] SerpAPI (ZipRecruiter + Glassdoor + Monster + Google Jobs)...")
    all_jobs += fetch_serpapi(seen)

    save_seen(seen)

    # Sort by: Role Priority → Company Tier → Has Salary
    all_jobs = sorted(all_jobs, key=get_sort_score)

    # Deduplicate by title+company
    seen_titles = set()
    unique_jobs = []
    for j in all_jobs:
        key = f"{j['title'].lower()}{j['company'].lower()}"
        if key not in seen_titles:
            seen_titles.add(key)
            unique_jobs.append(j)
    all_jobs = unique_jobs

    print(f"\nTotal new jobs found: {len(all_jobs)}")
    filepath = save_to_excel(all_jobs)
    send_summary(all_jobs)
    send_file(filepath)

# ── SCHEDULE ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    send_telegram(
        "🤖 <b>Full Job Bot is live!</b>\n"
        "⏰ Runs daily at 7:00 PM Cincinnati (EST)\n"
        "🔍 Sources: LinkedIn · Indeed · ZipRecruiter · Glassdoor · Monster\n"
        "         Adzuna · Jooble · RemoteOK · WeWorkRemotely · USAJobs\n"
        "         Greenhouse · Lever · Ashby · SmartRecruiters · Workday\n"
        "🏢 Covers 500+ top US company career pages via ATS\n"
        "📂 New Excel file daily: <b>Jobs_YYYY-MM-DD.xlsx</b>"
    )
    run()
    schedule.every().day.at("19:00").do(run)  # 7:00 PM Cincinnati EST
    while True:
        schedule.run_pending()
        time.sleep(60)
